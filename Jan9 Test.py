import pandas as pd
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import os


# Define start and end dates
start_date = datetime(2024, 1, 2)
end_date = datetime(2024, 1, 17)  # End date inclusive

# Define a list of pings and their corresponding filenames.
pings = ['ping1', 'ping2', 'ping3', 'ping4', 'ping5', 'ping6']
ping_files = [
    'ping1.csv',
    'ping2.csv',
    'ping3.csv',
    'ping4.csv',
    'ping5.csv',
    'ping6.csv'
]


# Generate a list of dates including the end date
date_generated = [start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]  # +1 makes inclusive

# Define the path to the Excel file
output_path = 'jan9main.xlsx'

# Check if the Excel file exists, if not, create it
if not os.path.exists(output_path):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        pd.DataFrame().to_excel(writer, index=False)  # Creating a default empty sheet

# Define a fill pattern for the header rows
header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


# Loop over each date to process data
for single_date in date_generated:
    all_rows = []  # Initialize list for all rows for the current date
    header_rows_index = []  # To keep track of header rows for coloring


# Process data for each ping
    for ping, ping_filename in zip(pings, ping_files):
        try:
            df = pd.read_csv(ping_filename)
        except FileNotFoundError:
            continue  # Skip if the file is not found

        # Adjust datetime to PST
        df['actual_start'] = pd.to_datetime(df['actual_start'])
        df['actual_start'] = df['actual_start'] - pd.Timedelta(hours=16)

        # Filter the DataFrame for the specific date
        filtered_df = df[df['actual_start'].dt.date == single_date.date()]

        if not filtered_df.empty:
            # Add an indicator column
            filtered_df.insert(0, 'Indicator', ping)

            # Mark the index where the header will be added
            header_rows_index.append(len(all_rows) + 1)  # +1 accounts for the header and the 1-index of Excel

            # Append the single row of data
            all_rows.append(filtered_df.columns.tolist())  # Add header
            all_rows.append(filtered_df.iloc[0].tolist())  # Add the single row of data

    # Convert the list of rows to a DataFrame
    main_df = pd.DataFrame(all_rows)

    # Define the sheet name for the current date
    sheet_name = single_date.strftime("%m-%d")

    # Write data to the Excel file
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
        book = writer.book
        # Check if the sheet exists and remove it
        if sheet_name in book.sheetnames:
            std = book[sheet_name]
            book.remove(std)
        main_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    # Apply the pattern fill to the header rows
    sheet = book[sheet_name]
    for header_row in header_rows_index:
        for cell in sheet[header_row]:
            cell.fill = header_fill

    # Save the Excel file after processing each date
    book.save(output_path)
